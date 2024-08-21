import os

from openpyxl import Workbook
import xlrd

def createComparison():
    wb = Workbook()

    newList = []
    """
    [(Index, ItemNum, Qt, REF),
    (Index, ItemNum, Qt, REF),
    (Index, ItemNum, Qt, REF)
    ]    
    """

    oldList = []
    """
    [(ItemNum, Qt),
    (ItemNum, Qt),
    (ItemNum, Qt)
    ]    
    """

    sheetOld = xlrd.open_workbook('old.xls', encoding_override = 'iso-8859-1').sheet_by_index(0)
    sheetNew = xlrd.open_workbook('new.xls', 'r').sheet_by_index(0)

    for row in range(0, sheetNew.nrows)[::-1]:
        indexStr = str(sheetNew.cell(row, 0).value).strip()
        itemNum = sheetNew.cell(row, 3).value.strip()
        quantityStr = str(sheetNew.cell(row, 5).value).strip()

        reference = str(sheetNew.cell(row, 4).value).strip() != ""

        index = indexStr.split('.')[0]
        if index.isdigit():
            index = int(index)
        else:
            index = -1

        quantity = quantityStr.split('.')[0]
        if quantity.isdigit():
            quantity = int(quantity)

        if index != -1:
            newList.append((index, itemNum, quantity, reference))



    for row in range(0, sheetOld.nrows)[::-1]:
        partNum = sheetOld.cell(row, 1).value.strip()
        valid = len(partNum.split('-')) == 3 and len(partNum) == 12

        quantityStr = str(sheetOld.cell(row, 8).value).strip()

        quantity = str(quantityStr.split('.')[0])

        if quantity.isdigit():
            quantity = int(quantity)

        if valid:
            oldList.append((partNum, quantity))


    sheetCompare = wb.create_sheet("Compare", 0)

    unfoundItems = []
    changes = []

    for part in newList:
        if part[0] != '':
            row = str(part[0])

            if not part[3]:
                sheetCompare['A' + row] = part[0]
                sheetCompare['B' + row] = part[1]
                sheetCompare['C' + row] = part[2]

            foundItem = False
            for item in oldList:
                if item[0] == part[1] and not foundItem:
                    foundItem = True

                    difference = 0
                    if part[2] != "" and item[1] != "":
                        difference = int(part[2]) - int(item[1])
                    elif part[2] == "" and item[1] == "":
                        difference = 0
                    elif part[2] != "" and item[1] == "":
                        difference = int(part[2])
                    elif part[2] == "" and item[1] != "":
                        difference = int(item[1])

                    if not part[3]:
                        sheetCompare['E' + row] = item[0]
                        sheetCompare['F' + row] = item[1]

                        if difference:
                            sheetCompare['H' + row] = difference

                            changes.append((part[0], part[1], difference))

                    oldList.pop(oldList.index(item))

            if not part[3]:
                if not foundItem:
                    unfoundItems.append((part[0], part[1], part[2]))
                    changes.append((part[0], part[1], part[2]))

                    sheetCompare['H' + row] = part[2]

    infoRow = max(len(sheetCompare['A']), len(sheetCompare['E'])) + 6

    sheetCompare['B' + str(infoRow-1)] = "Unmatched Items from AutoCAD:"
    sheetCompare['A' + str(infoRow)] = "Item #:"
    sheetCompare['B' + str(infoRow)] = "Manufacture Num.:"
    sheetCompare['D' + str(infoRow)] = "Quantity:"
    infoRow += 1

    for item in unfoundItems:
        sheetCompare['A' + str(infoRow)] = item[0]
        sheetCompare['B' + str(infoRow)] = item[1]
        sheetCompare['C' + str(infoRow)] = item[1]
        infoRow += 1

    infoRow += 3

    sheetCompare['B' + str(infoRow-1)] = "Unmatched Items from the DBA:"
    sheetCompare['E' + str(infoRow)] = "Manufacture Num.:"
    sheetCompare['F' + str(infoRow)] = "Quantity:"
    infoRow += 1

    for item in reversed(oldList):
        sheetCompare['E' + str(infoRow)] = item[0]
        sheetCompare['F' + str(infoRow)] = item[1]
        infoRow += 1

    infoRow += 3
    sheetCompare['C' + str(infoRow-1)] = "Changed Items:"
    sheetCompare['A' + str(infoRow)] = "Item #:"
    sheetCompare['B' + str(infoRow)] = "Manufacture Num.:"
    sheetCompare['D' + str(infoRow)] = "Qty Diff.:"
    infoRow += 1

    for item in reversed(changes):
        sheetCompare['A' + str(infoRow)] = item[0]
        sheetCompare['B' + str(infoRow)] = item[1]
        sheetCompare['D' + str(infoRow)] = item[2]
        infoRow += 1

    wb.save("Compare.xlsx")


try:
    createComparison()
    print("Created/Updated file: 'Compare.xlsx'")
    input("Press Enter to open file...")
    os.startfile('Compare.xlsx')

except:
    print("Error: Make sure you have renamed your files (AutoCAD export as 'new.xls' and DBA export as 'old.xls') and that you have closed any open 'comparison.xlsx' files from previous runs.")
    input("Press Enter to exit...")
